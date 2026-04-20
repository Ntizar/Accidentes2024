from __future__ import annotations

import json
import re
import unicodedata
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "Accidentes-con-victimas-Tablas-estadisticas-2024.xlsx"
PUBLIC_DATA_PATH = ROOT / "public" / "data"
PROVINCES_GEOJSON_PATH = PUBLIC_DATA_PATH / "spain-provinces.geojson"
PROVINCES_GEOJSON_URL = (
    "https://raw.githubusercontent.com/codeforgermany/"
    "click_that_hood/main/public/data/spain-provinces.geojson"
)

PROVINCE_ALIASES = {
    "balears illes": "illes balears",
    "coruna a": "a coruna",
    "rioja la": "la rioja",
    "palmas las": "las palmas",
    "alicante alacant": "alacant alicante",
    "castellon castello": "castello castellon",
    "bizkaia": "bizkaia vizcaya",
    "gipuzkoa": "gipuzkoa guipuzcoa",
}

ROAD_LABELS = {
    "motorway": "Autopista",
    "dual_carriageway": "Autovia",
    "conventional": "Via convencional",
    "rural_lane": "Camino vecinal",
    "service_road": "Via de servicio",
    "link_road": "Ramal de enlace",
    "other": "Otro tipo",
    "street": "Calle",
    "crossing": "Travesia",
}


def normalize_text(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, float) and value.is_integer():
        return int(value)

    if isinstance(value, str):
        compact = re.sub(r"\s+", " ", value.replace("\n", " ")).strip()
        return compact or None

    return value


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def to_int(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).strip())


def slugify(value):
    normalized = normalize_text(value) or ""
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower()
    normalized = normalized.replace("/", " ")
    normalized = normalized.replace("(", " ").replace(")", " ")
    normalized = re.sub(r"[^a-z0-9 ]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return PROVINCE_ALIASES.get(normalized, normalized)


def sheet_id(title):
    raw = slugify(title)
    return raw.replace(" ", "-")


def expand_sheet(ws):
    merge_map = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merge_map[(row, col)] = top_left

    rows = []
    last_row = 0
    last_col = 0

    for row_index in range(1, ws.max_row + 1):
        row_values = []
        for col_index in range(1, ws.max_column + 1):
            value = ws.cell(row_index, col_index).value
            if value is None and (row_index, col_index) in merge_map:
                value = merge_map[(row_index, col_index)]

            value = normalize_text(value)
            row_values.append(value)

            if value is not None:
                last_row = max(last_row, row_index)
                last_col = max(last_col, col_index)

        rows.append(row_values)

    trimmed = [row[:last_col] for row in rows[:last_row]]
    return trimmed


def find_data_start(rows):
    for index in range(2, len(rows)):
        current = rows[index]
        previous = rows[index - 1]
        has_label = len(current) > 0 and current[0] is not None
        has_numbers = any(is_number(value) for value in current[1:])
        previous_has_numbers = any(is_number(value) for value in previous[1:])

        if has_label and has_numbers and not previous_has_numbers:
            return index

    for index, row in enumerate(rows):
        if row and row[0] is not None and any(is_number(value) for value in row[1:]):
            return index

    return len(rows)


def build_split_metrics(row, start_index):
    return {
        "accidents": to_int(row[start_index]),
        "fatalAccidents": to_int(row[start_index + 1]),
        "fatalities": to_int(row[start_index + 2]),
        "hospitalized": to_int(row[start_index + 3]),
        "nonHospitalized": to_int(row[start_index + 4]),
    }


def build_time_metrics(row, start_index):
    return {
        "accidents": to_int(row[start_index]),
        "fatalAccidents30Days": to_int(row[start_index + 1]),
        "victims": to_int(row[start_index + 2]),
        "fatalities": to_int(row[start_index + 3]),
        "hospitalized": to_int(row[start_index + 4]),
        "nonHospitalized": to_int(row[start_index + 5]),
    }


def build_road_metrics(row, start_index):
    return {
        "accidents": to_int(row[start_index]),
        "fatalities": to_int(row[start_index + 1]),
        "hospitalized": to_int(row[start_index + 2]),
        "nonHospitalized": to_int(row[start_index + 3]),
    }


def parse_split_table(rows, key_name):
    data_start = find_data_start(rows)
    records = []
    total = None

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None:
            continue

        record = {
            key_name: label,
            "slug": slugify(label),
            "interurban": build_split_metrics(row, 1),
            "urban": build_split_metrics(row, 6),
            "total": build_split_metrics(row, 11),
        }

        if record["slug"] == "total":
            total = record
        else:
            records.append(record)

    return records, total


def parse_time_series(rows, key_name):
    data_start = find_data_start(rows)
    records = []
    total = None

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None:
            continue

        record = {
            key_name: label,
            "slug": slugify(label),
            "interurban": build_time_metrics(row, 1),
            "urban": build_time_metrics(row, 7),
            "total": build_time_metrics(row, 13),
        }

        if record["slug"] == "total":
            total = record
        else:
            records.append(record)

    return records, total


def parse_hourly(rows):
    data_start = find_data_start(rows)
    weekdays = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
    records = []

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None or slugify(label) == "total":
            continue

        day_values = {weekday: to_int(row[index + 1]) for index, weekday in enumerate(weekdays)}
        record = {
            "hour": label,
            **day_values,
            "total": to_int(row[8]),
        }
        records.append(record)

    return records


def parse_user_classes(rows):
    data_start = find_data_start(rows)
    roles = {
        "total": 1,
        "driver": 6,
        "passenger": 11,
        "pedestrian": 16,
    }
    records = []

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None or slugify(label) == "total":
            continue

        role_data = {}
        for role_name, start_index in roles.items():
            role_data[role_name] = {
                "implicated": to_int(row[start_index]),
                "victims": to_int(row[start_index + 1]),
                "fatalities": to_int(row[start_index + 2]),
                "hospitalized": to_int(row[start_index + 3]),
                "nonHospitalized": to_int(row[start_index + 4]),
            }

        records.append(
            {
                "className": label,
                "slug": slugify(label),
                **role_data,
            }
        )

    return records


def parse_age_distribution(rows):
    data_start = find_data_start(rows)
    grouped = []
    current_age = None
    lookup = {}

    for row in rows[data_start:]:
        age_label = normalize_text(row[0]) or current_age
        sex_label = normalize_text(row[1])
        current_age = age_label

        if age_label is None or sex_label is None:
            continue

        age_slug = slugify(age_label)
        if age_slug == "total":
            continue

        if age_slug not in lookup:
            lookup[age_slug] = {
                "ageGroup": age_label,
                "slug": age_slug,
                "male": {"fatalities": 0, "hospitalized": 0, "nonHospitalized": 0, "totalVictims": 0},
                "female": {"fatalities": 0, "hospitalized": 0, "nonHospitalized": 0, "totalVictims": 0},
                "unknown": {"fatalities": 0, "hospitalized": 0, "nonHospitalized": 0, "totalVictims": 0},
                "total": {"fatalities": 0, "hospitalized": 0, "nonHospitalized": 0, "totalVictims": 0},
            }
            grouped.append(lookup[age_slug])

        target = lookup[age_slug]
        segment_key = {
            "hombre": "male",
            "mujer": "female",
            "se desconoce": "unknown",
            "total": "total",
        }.get(slugify(sex_label))

        if segment_key is None:
            continue

        segment = {
            "fatalities": to_int(row[2]),
            "hospitalized": to_int(row[3]),
            "nonHospitalized": to_int(row[4]),
        }
        segment["totalVictims"] = (
            segment["fatalities"] + segment["hospitalized"] + segment["nonHospitalized"]
        )
        target[segment_key] = segment

    return grouped


def parse_driver_infractions(rows):
    data_start = find_data_start(rows)
    vehicle_keys = [
        "bicycle",
        "vmp",
        "cyclomotor",
        "motorcycle",
        "tourism",
        "van",
        "truckUpTo3500",
    ]
    records = []

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None or slugify(label) == "total":
            continue

        entry = {
            "infraction": label,
            "slug": slugify(label),
            "total": to_int(row[1]),
        }

        for index, vehicle_key in enumerate(vehicle_keys, start=2):
            entry[vehicle_key] = to_int(row[index])

        records.append(entry)

    return records


def parse_road_types(rows):
    data_start = find_data_start(rows)
    interurban_types = [
        ("motorway", 1),
        ("dual_carriageway", 5),
        ("conventional", 9),
        ("rural_lane", 13),
        ("service_road", 17),
        ("link_road", 21),
        ("other", 25),
    ]
    urban_types = [("street", 33), ("crossing", 37)]
    records = []

    for row in rows[data_start:]:
        label = normalize_text(row[0])
        if label is None or slugify(label) == "total":
            continue

        interurban = {key: build_road_metrics(row, offset) for key, offset in interurban_types}
        urban = {key: build_road_metrics(row, offset) for key, offset in urban_types}

        dominant_interurban_key = max(interurban, key=lambda key: interurban[key]["accidents"])
        dominant_urban_key = max(urban, key=lambda key: urban[key]["accidents"])

        records.append(
            {
                "name": label,
                "slug": slugify(label),
                "interurban": {
                    "roadTypes": interurban,
                    "total": build_road_metrics(row, 29),
                },
                "urban": {
                    "roadTypes": urban,
                    "total": build_road_metrics(row, 41),
                },
                "dominantInterurbanRoadType": {
                    "key": dominant_interurban_key,
                    "label": ROAD_LABELS[dominant_interurban_key],
                    "accidents": interurban[dominant_interurban_key]["accidents"],
                },
                "dominantUrbanRoadType": {
                    "key": dominant_urban_key,
                    "label": ROAD_LABELS[dominant_urban_key],
                    "accidents": urban[dominant_urban_key]["accidents"],
                },
            }
        )

    return records


def fetch_provinces_geojson():
    try:
        with urllib.request.urlopen(PROVINCES_GEOJSON_URL) as response:
            return json.load(response)
    except Exception:
        if PROVINCES_GEOJSON_PATH.exists():
            return json.loads(PROVINCES_GEOJSON_PATH.read_text(encoding="utf-8"))
        raise


def enrich_geojson(geojson, provinces, road_types):
    province_lookup = {province["slug"]: province for province in provinces}
    road_lookup = {road["slug"]: road for road in road_types}
    unmatched = []

    for feature in geojson.get("features", []):
        feature_slug = slugify(feature.get("properties", {}).get("name"))
        province = province_lookup.get(feature_slug)
        road = road_lookup.get(feature_slug)

        if province is None:
            unmatched.append(feature.get("properties", {}).get("name"))
            continue

        feature.setdefault("properties", {})
        feature["properties"].update(
            {
                "dashboardKey": province["slug"],
                "dashboardName": province["name"],
                "totalAccidents": province["total"]["accidents"],
                "totalFatalities": province["total"]["fatalities"],
                "totalHospitalized": province["total"]["hospitalized"],
                "totalNonHospitalized": province["total"]["nonHospitalized"],
                "urbanAccidents": province["urban"]["accidents"],
                "interurbanAccidents": province["interurban"]["accidents"],
                "urbanFatalities": province["urban"]["fatalities"],
                "interurbanFatalities": province["interurban"]["fatalities"],
            }
        )

        if road is not None:
            feature["properties"].update(
                {
                    "dominantInterurbanRoadType": road["dominantInterurbanRoadType"]["label"],
                    "dominantUrbanRoadType": road["dominantUrbanRoadType"]["label"],
                }
            )

    if unmatched:
        print("GeoJSON features without workbook match:", ", ".join(unmatched))

    return geojson


def build_table_exports(sheets):
    exported = []

    for title, rows in sheets.items():
        data_start = find_data_start(rows)
        header_rows = rows[2:data_start]
        body_rows = rows[data_start:]
        column_count = max((len(row) for row in rows), default=0)

        exported.append(
            {
                "sheetId": sheet_id(title),
                "sheetName": title,
                "title": title,
                "description": normalize_text(rows[0][0]) if rows else title,
                "headerRows": [[cell or "" for cell in row] for row in header_rows],
                "rows": [[cell or "" for cell in row] for row in body_rows],
                "rowCount": len(body_rows),
                "columnCount": column_count,
            }
        )

    return exported


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    PUBLIC_DATA_PATH.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)
    sheets = {worksheet.title: expand_sheet(worksheet) for worksheet in workbook.worksheets}

    provinces, national_totals = parse_split_table(sheets["TABLA 1.1"], "name")
    communities, national_communities_total = parse_split_table(
        sheets["TABLA 1.1.C.A."], "name"
    )
    accident_types, accident_types_total = parse_split_table(sheets["TABLA 1.3"], "type")
    months, months_total = parse_time_series(sheets["TABLA 3.1"], "month")
    weekdays, weekdays_total = parse_time_series(sheets["TABLA 3.2"], "weekday")

    hourly = {
        "interurban": parse_hourly(sheets["TABLA 3.4.I"]),
        "urban": parse_hourly(sheets["TABLA 3.4.U"]),
    }

    user_classes = {
        "interurban": parse_user_classes(sheets["TABLA 2.2.I"]),
        "urban": parse_user_classes(sheets["TABLA 2.2.U"]),
    }

    age_distribution = {
        "interurban": parse_age_distribution(sheets["TABLA 4.1.I"]),
        "urban": parse_age_distribution(sheets["TABLA 4.1.U"]),
    }

    driver_infractions = {
        "interurban": parse_driver_infractions(sheets["TABLA 6.1.I"]),
        "urban": parse_driver_infractions(sheets["TABLA 6.1.U"]),
    }

    road_types_by_province = parse_road_types(sheets["TABLA 7.1"])

    provinces_geojson = enrich_geojson(
        fetch_provinces_geojson(),
        provinces,
        road_types_by_province,
    )

    tables_export = build_table_exports(sheets)
    available_tables = [
        {
            "sheetId": table["sheetId"],
            "sheetName": table["sheetName"],
            "description": table["description"],
            "rowCount": table["rowCount"],
            "columnCount": table["columnCount"],
        }
        for table in tables_export
    ]

    top_accidents = sorted(provinces, key=lambda item: item["total"]["accidents"], reverse=True)[:12]
    top_fatalities = sorted(provinces, key=lambda item: item["total"]["fatalities"], reverse=True)[:12]
    top_communities = sorted(communities, key=lambda item: item["total"]["accidents"], reverse=True)[:8]
    peak_month = max(months, key=lambda item: item["total"]["accidents"])
    peak_weekday = max(weekdays, key=lambda item: item["total"]["accidents"])

    dashboard = {
        "meta": {
            "title": "Accidentes con victimas 2024",
            "year": 2024,
            "sourceFile": WORKBOOK_PATH.name,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "note": (
                "El Excel aporta datos agregados por provincia, comunidad y categoria; "
                "no incluye siniestros individuales geolocalizados punto a punto."
            ),
        },
        "overview": {
            "national": national_totals,
            "nationalCommunitiesTotal": national_communities_total,
            "provinces": provinces,
            "communities": communities,
            "topProvincesByAccidents": top_accidents,
            "topProvincesByFatalities": top_fatalities,
            "topCommunitiesByAccidents": top_communities,
            "defaultProvince": top_accidents[0]["slug"] if top_accidents else None,
        },
        "trends": {
            "months": months,
            "monthsTotal": months_total,
            "weekdays": weekdays,
            "weekdaysTotal": weekdays_total,
            "hourly": hourly,
            "peakMonth": {"month": peak_month["month"], "accidents": peak_month["total"]["accidents"]},
            "peakWeekday": {
                "weekday": peak_weekday["weekday"],
                "accidents": peak_weekday["total"]["accidents"],
            },
        },
        "victims": {
            "userClasses": user_classes,
            "ageDistribution": age_distribution,
        },
        "risk": {
            "accidentTypes": accident_types,
            "accidentTypesTotal": accident_types_total,
            "driverInfractions": driver_infractions,
            "roadTypesByProvince": road_types_by_province,
        },
        "tables": {
            "available": available_tables,
        },
    }

    tables_payload = {
        "meta": dashboard["meta"],
        "sheets": tables_export,
    }

    dashboard_path = PUBLIC_DATA_PATH / "dashboard.json"
    tables_path = PUBLIC_DATA_PATH / "tables.json"

    dashboard_path.write_text(json.dumps(dashboard, ensure_ascii=False, indent=2), encoding="utf-8")
    tables_path.write_text(json.dumps(tables_payload, ensure_ascii=False), encoding="utf-8")
    PROVINCES_GEOJSON_PATH.write_text(
        json.dumps(provinces_geojson, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Generated {dashboard_path}")
    print(f"Generated {tables_path}")
    print(f"Generated {PROVINCES_GEOJSON_PATH}")


if __name__ == "__main__":
    main()
